import requests
from bs4 import BeautifulSoup as bs
import openpyxl
from openpyxl.styles import Alignment,fonts,NamedStyle
import os
import telebot 
from Alive import alive
#########
alive()

BOT_TOKEN =os.environ.get('token')

bot = telebot.TeleBot(BOT_TOKEN)


#############
centered_style = NamedStyle(name="centered_style", alignment=Alignment(horizontal='center', vertical='center'))
start_style = NamedStyle(name="centered_style", alignment=Alignment(horizontal='right', vertical='center'))
  
e=4
r=3
w=2
wb=openpyxl.Workbook()


ws=wb.active
recent=wb.create_sheet("Recent")
stats=wb.create_sheet("stats")

recent.freeze_panes='B1'
stats.freeze_panes='C1'


main_url="https://www.espncricinfo.com"
#test_id='/player/shadab-khan-922943'
header = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.76 Safari/537.36'}
	

def number(d):
	try:
		wc=int(d)
	except:
		wc=d
	#print (type(d))
	return wc

def excel_name(url):
	global name
	link=url+"/live-cricket-score"
	response=requests.get(link )
	if response.status_code==200:
		soup=bs(response.content,"html.parser")
		data_text=soup.find('h1').text
		data_text_list=data_text.split(",")
		match_date=data_text_list[-1].split('-')[0]
		match_date=match_date[1:-6]
		name=f'{data_text_list[0]} ({match_date}).xlsx'
		
		return name
			
def player_details(p_id):
	link2=main_url+p_id
	#print(link2)
	response=requests.get(link2,header)
	if response.status_code==200:
		soup=bs(response.content,"html.parser")
			#print(soup)
		darkari=soup.find('div', class_= "ds-grid lg:ds-grid-cols-3 ds-grid-cols-2 ds-gap-4 ds-mb-8")
		nirdistos=darkari.find_all('span', class_="ds-text-title-s ds-font-bold ds-text-typo")
			
		player=[]
		for nirdisto in nirdistos:
				data=nirdisto.find("p")
				#print(data)
				try:
					player.append(data.string)
					pass
				except:
					pass
		
		#print(player)
		return player



def player_details_data(team_data):
	team=dict()
	for player in team_data:
	       try:
	        data=player.find('a')
	        player_param=data.get('href')
	        player_name=player.find('span'). string 
	        player_type=player.find('p'). string 
	        team[player_name]=player_param
	       except:
	       	pass
	       #print(team)
	return team

def get_player(link):
	team1_data=[ ]
	team2_data=[ ]
	k=-1
	url=link+"/match-squads"
	response=requests.get(url, header)
	   # print(response.status_code)
	if response.status_code==200:
	        soup=bs(response.content, 'html.parser')
	        
	       
	        finder=soup.find_all('td', class_='ds-min-w-max')
	    
	        while k != len(finder)-1:
	                     k +=1
	                     k+=1
	                     team1_data.append(finder[k])
	                     k+=1
	                     team2_data.append(finder[k])
	        	
	        team1=player_details_data(team1_data)
	        team2=player_details_data(team2_data)
	        #print(team1)
	        #print(team2)
	return team1,team2


#t=get_player(test_url)

#print(t)
#print("\n")
#print(t2)



	
def player_recents(p_id):
	recent_matches=[ ]
	url=main_url+p_id+'/matches'
	#print(url)
	response=requests.get(url,header)
	if response.status_code==200:
		#print('We finding')
		soup=bs(response.content, 'html.parser')
		table_rows=soup.find_all( "tr", class_="ds-text-tight-m")
		#print (len(table_rows))
		for table_row in table_rows:
			#print(table_row)
			table_data=table_row.find_all("td", class_="ds-min-w-max")
			#print("\n"*5)
			last_match=[ ]
			for t in table_data:
				last_match.append(t.string)
			recent_matches.append(last_match)
		#print(recent_matches)
	return recent_matches
	

	

	
def player_stats(m_id):
		stats=[]
		url=main_url+m_id
		#print(url)
		response=requests.get(url,header)
		if response.status_code==200:
			soup=bs(response.content, 'html.parser')
			first_search=soup.find_all("div", class_="ds-p-0")
			
			for a in first_search:
				if a.next_element.next_element.name == 'p':
					table_finder_soups=a.find_all("table")
					for table in  table_finder_soups:
						stat=[]
						table_headings=table.find_all("th")
						row_heading=[ ]
						for th in table_headings:
							row_heading.append(th.string)
						stat.append(row_heading)
						table_datas=table.find_all("tr")
						for table_data in table_datas:
							cell_data= table_data.find_all("td")
							row_data=[]
					
							for cell in cell_data:
								row_data.append(cell.string)
							#print(row_data)
							stat.append(row_data)
						#print("\n"*3)
						stats.append(stat)
						#print(stat)
		#print(stats)
		return stats






def write_my(m_id):
	global r,e
	n=player_details(m_id)
	recent.cell(row=r, column=1). value=n[0]
	for j in n[1:]:
		r+=1
		recent.cell(row=r, column=2). value=j
	
	r+=2
	for data_colum in player_recents(m_id):
			k=3
			for data in data_colum:
				recent.cell(row=r,column=k).value= number(data)
				recent.cell(row=r,column=k).style=centered_style

				k+=1
			r+=1
	r+=4
	k=1
	stats.cell(row=e,column=1).value=n[0]
	e+=1
	statics=player_stats(m_id)
	for stat in statics:
		e+=1
		for row_line in stat:
			if len(row_line)== 0:
				continue
			k=1
			for p in row_line:
				stats.cell(row=e,column=k).value=number(p)
				stats.cell(row=e,column=k).style=centered_style
				k+=1
			e+=1
	e+=4
	recent.column_dimensions['C'].width=25
	recent.column_dimensions['E'].width=10
	recent.column_dimensions['F'].width=12
	recent.column_dimensions['G'].width=13

    
    
	wb.save(name)
	print(n)
	return 
	
def ground_link(today):
	url=today+'/live-cricket-score'
	#print(url)
	response=requests.get(url,header)
	if response.status_code==200:
		soup=bs(response.content, 'html.parser')
		ground_soup=soup.find("td",class_="ds-min-w-max",colspan="2")
		ground_link_soup=ground_soup.find("a")
		ground_name=ground_soup.text
		link=ground_link_soup.get("href")
		#print (link)
		return link



def get_groundwise_all_suffix(u):
	global w
	typewise_odit=[]
	my_need=[]
	response=requests.get(u,header)
	if response.status_code==200:
			soup=bs(response.content, 'html.parser')
			main_box=soup.find('div', id='recs')
			data_title=main_box.find('h2'). string 

			#print(data_title)
			p=w-1
			ws.cell(row=p,column=3).value=data_title
			w+=1
			datas=main_box.find_all("tr", class_='islast1')
			typewise_odit.append(datas[1])
			typewise_odit.append(datas[2])
			''' typewise_dict={}
			for data in datas:
				match_typewise_datas=data.find_all('a')
				typename=match_typewise_datas[0].text
				test=data[0].
				typewise_dict[typename]=link '''
			#print(typewise_odit)
			for type in typewise_odit:
				#print (type)
				ground_typewise_datas=type.find_all('a')
				my_need.append(ground_typewise_datas[2])
				my_need.append(ground_typewise_datas[3])
				my_need.append(ground_typewise_datas[5])
				my_need.append(ground_typewise_datas[6])
	my_all_suffix=[ ]
	for x in my_need:
		suffix_link=x.get("href")
		my_all_suffix.append(suffix_link)
	#print(my_all_suffix)
	return my_all_suffix


def statistics_table_maker(suffix):
	global w
	type_of_match=["ODI","T20"]
	#e_url="https://stats.espncricinfo.com/ci/engine/records/team/highest_innings_totals.html?class=2;id=683;type=ground"
	page_url="https://stats.espncricinfo.com"+suffix
	#print(page_url)
	response=requests.get(page_url)
	d=1
	if response.status_code==200:
		soup=bs(response.content, 'html.parser')
		table_soup=soup.find('table', class_="engineTable")
		title=table_soup.find("caption"). text
		print(title)
		ws.cell(row=w,column=4).value=title
		w+=1
		table_heading=[ ]
		table_heads_soup=table_soup.find_all("th")
		for table_head in table_heads_soup:
			#print (table_head.text)
			table_heading.append(table_head.text)
			ws.cell(row=w,column=d).value=table_head.text
			d=d+1
		#print(table_heading,"\n")
		
		table_row_soups=table_soup.find_all("tr")
		for table_row_soup in table_row_soups:
			table_data_soups=table_row_soup.find_all("td")
			
			d=1
			for table_data_soup in table_data_soups:
				table_data=table_data_soup.text
				ws.cell(row=w,column=d).value=table_data
				d+=1
			w+=1
				#print(table_data)
	w+=2
	ws.title="Ground"
	wb.save(name)
	return
	


def extract_data(dict):
		for p in dict:
			g=f"{dict[p]}"
			write_my(g)
			print('Doing')
		return 
		
#g_url=ground_link(today_url)
'''count=0
for x in get_groundwise_all_suffix(g_url):
	if count==0:
		ws.cell(row=w,column=4).value="ODI"
		w+=1
	if count==4:
		w+=4
		ws.cell(row=w,column=4).value="T20"
		w+=1
	statistics_table_maker(x)
	count+=1	'''
	
	
	
def final(url):
	today_url=url
	if "match-previe"  in today_url:
		today_url=today_url.replace("/match-preview","")
	elif "live-cricket-score"  in today_url:
		today_url=today_url.replace("/live-cricket-score","")
	else:
		pass
	excel_name(today_url)

	for a in get_player(today_url):
		extract_data(a)
	print("Complete")
	return name



############ Bot section #########
@bot.message_handler(commands=['start' , 'Start'])
def start(message):
    bot.reply_to(message, "Hi! I can provide Excel files of any Criket match from ESPNcricinfo website\n"
                          "Enter the corresponding match link.")
@bot.message_handler(func=lambda message: True)
def handle_link(message):
    provide_link = message.text
    bot.send_message(message.chat.id, "wait we are preparing Excel file")
    name=final(provide_link)
    bot.send_document(message.chat.id, name, caption='Here is your Excel file!')
    bot.send_message(message.chat.id, "done")
    
def handle_excel_request(message):
  user_data = message.text.strip() 
  


bot.polling()


